import csv
import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


def _normalize_row(row, keys):
    if isinstance(row, dict):
        return [row.get(key, '') for key in keys]
    return [getattr(row, key, '') for key in keys]


def export_csv(filename, columns, keys, rows):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(_normalize_row(row, keys))
    return response


def export_xlsx(filename, columns, keys, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = filename[:31]
    sheet.append(columns)
    for row in rows:
        sheet.append(_normalize_row(row, keys))

    for index, _ in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].auto_size = True

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_pdf(title, filename, columns, keys, rows):
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles['Heading2']), Spacer(1, 12)]

    table_data = [columns]
    for row in rows:
        table_data.append([str(value) for value in _normalize_row(row, keys)])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]
        )
    )
    story.append(table)
    document.build(story)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
